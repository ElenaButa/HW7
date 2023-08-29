from openpyxl import load_workbook
import os.path
from conftest import resources_path

# TODO оформить в тест, добавить ассерты и использовать универсальный путь



def test_file():
    workbook = load_workbook(os.path.join(resources_path,"file_example_XLSX_50.xlsx"))
    sheet = workbook.active
    print(sheet.cell(row=2, column=3).value)
    assert (sheet.cell(row=3, column=2).value)=='Mara'
    assert (sheet.cell(row=6, column=2).value)=='Nereida'